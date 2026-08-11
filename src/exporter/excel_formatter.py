from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.utils import get_column_letter


class ExcelFormatter:

    def format(self, worksheet):

        # ----------------------------------
        # Header Style
        # ----------------------------------

        header_fill = PatternFill(
            fill_type="solid",
            start_color="1F4E78",
            end_color="1F4E78"
        )

        header_font = Font(
            bold=True,
            color="FFFFFF",
            size=11
        )

        thin_border = Border(

            left=Side(style="thin"),

            right=Side(style="thin"),

            top=Side(style="thin"),

            bottom=Side(style="thin")

        )

        for cell in worksheet[1]:

            cell.fill = header_fill

            cell.font = header_font

            cell.border = thin_border

            cell.alignment = Alignment(

                horizontal="center",

                vertical="center",

                wrap_text=True

            )

        # ----------------------------------
        # Freeze Header
        # ----------------------------------

        worksheet.freeze_panes = "A2"

        # ----------------------------------
        # Auto Filter
        # ----------------------------------

        worksheet.auto_filter.ref = worksheet.dimensions

        # ----------------------------------
        # Alignment + Borders
        # ----------------------------------

        numeric_headers = {

            "File Size (Bytes)",

            "Word Count",

            "Token Count",

            "Character Count",

            "Reading Time (Minutes)"

        }

        for row in worksheet.iter_rows(min_row=2):

            for cell in row:

                header = worksheet.cell(
                    row=1,
                    column=cell.column
                ).value

                if header in numeric_headers:

                    cell.alignment = Alignment(

                        horizontal="center",

                        vertical="top"

                    )

                else:

                    cell.alignment = Alignment(

                        vertical="top",

                        wrap_text=True

                    )

                cell.border = thin_border

        # ----------------------------------
        # Preferred Widths
        # ----------------------------------

        preferred_widths = {

            "Title": 40,

            "Source": 18,

            "Source Type": 15,

            "Source URL": 50,

            "Canonical URL": 50,

            "Author": 25,

            "Publisher": 30,

            "Language": 15,

            "Language Code": 12,

            "Domain": 28,

            "Subdomain": 30,

            "Description": 70,

            "Keywords": 45,

            "Published Date": 20,

            "Modified Date": 20,

            "File Name": 40,

            "File Extension": 15,

            "File Size (Bytes)": 18,

            "Word Count": 15,

            "Token Count": 15,

            "Character Count": 18,

            "Reading Time (Minutes)": 22,

            "Processed At": 25,

            "Pipeline Version": 22

        }

        for column in worksheet.columns:

            header = column[0].value

            column_letter = get_column_letter(
                column[0].column
            )

            if header in preferred_widths:

                worksheet.column_dimensions[
                    column_letter
                ].width = preferred_widths[
                    header
                ]

            else:

                max_length = 0

                for cell in column:

                    try:

                        if cell.value:

                            max_length = max(
                                max_length,
                                len(str(cell.value))
                            )
                    except Exception:

                        pass

                worksheet.column_dimensions[
                    column_letter
                ].width = min(
                    max_length + 3,
                    50
                )

        # ----------------------------------
        # Header Height
        # ----------------------------------

        worksheet.row_dimensions[1].height = 30

        return worksheet