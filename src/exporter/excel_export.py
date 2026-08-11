from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.exporter.excel_formatter import ExcelFormatter


class ExcelExporter:

    def __init__(self):

        self.output_dir = Path("metadata/excel")

        self.output_dir.mkdir(
            parents=True,
            exist_ok=True
        )

        self.file = self.output_dir / "metadata.xlsx"

        self.formatter = ExcelFormatter()

        self.headers = [

            "Title",

            "Source",

            "Source Type",

            "Source URL",

            "Canonical URL",

            "Author",

            "Publisher",

            "Language",

            "Language Code",

            "Domain",

            "Subdomain",

            "Description",

            "Keywords",

            "Published Date",

            "Modified Date",

            "File Name",

            "File Extension",

            "File Size (Bytes)",

            # -------------------------
            # YouTube Metadata
            # -------------------------

            "Channel",

            "Channel ID",

            "Upload Date",

            "Duration (Seconds)",

            "Thumbnail",

            "Tags",

            "Categories",

            "View Count",

            "Like Count",

            "Comment Count",

            # -------------------------
            # Statistics
            # -------------------------

            "Word Count",

            "Token Count",

            "Character Count",

            "Reading Time (Minutes)",

            # -------------------------
            # Pipeline
            # -------------------------

            "Processed At",

            "Pipeline Version"

        ]

        if not self.file.exists():

            wb = Workbook()

            ws = wb.active

            ws.title = "Metadata"

            ws.append(self.headers)

            self.formatter.format(ws)

            wb.save(self.file)

    # -------------------------------------------------

    def export(self, metadata):

        wb = load_workbook(self.file)

        ws = wb.active

        ws.append([

            metadata.get("title", ""),

            metadata.get("source", ""),

            metadata.get("source_type", ""),

            metadata.get("source_url", ""),

            metadata.get("canonical_url", ""),

            metadata.get("author", ""),

            metadata.get("publisher", ""),

            metadata.get("language", ""),

            metadata.get("language_code", ""),

            metadata.get("domain", ""),

            metadata.get("subdomain", ""),

            metadata.get("description", ""),

            metadata.get("keywords", ""),

            metadata.get("published_date", ""),

            metadata.get("modified_date", ""),

            metadata.get("file_name", ""),

            metadata.get("file_extension", ""),

            metadata.get("file_size_bytes", 0),

            # -------------------------
            # YouTube Metadata
            # -------------------------

            metadata.get("channel", ""),

            metadata.get("channel_id", ""),

            metadata.get("upload_date", ""),

            metadata.get("duration_seconds", 0),

            metadata.get("thumbnail", ""),

            metadata.get("tags", ""),

            metadata.get("categories", ""),

            metadata.get("view_count", 0),

            metadata.get("like_count", 0),

            metadata.get("comment_count", 0),

            # -------------------------
            # Statistics
            # -------------------------

            metadata.get("word_count", 0),

            metadata.get("token_count", 0),

            metadata.get("character_count", 0),

            metadata.get("reading_time_minutes", 0),

            # -------------------------
            # Pipeline
            # -------------------------

            metadata.get("processed_at", ""),

            metadata.get("pipeline_version", "")

        ])

        self.formatter.format(ws)

        wb.save(self.file)

        return self.file